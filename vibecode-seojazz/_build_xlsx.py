#!/usr/bin/env python3
# Строит struktura-konkurenty-url.xlsx из исходной структуры + _konkurenty.json
import json, glob, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
src = None
for p in ['/sessions/affectionate-upbeat-planck/mnt/uploads/struktura-seojazz.xlsx']:
    if os.path.exists(p): src = p
if not src:
    cand = glob.glob('/sessions/**/struktura-seojazz.xlsx', recursive=True)
    src = cand[0] if cand else None

data = json.load(open(os.path.join(HERE, '_konkurenty.json'), encoding='utf-8'))
wb = openpyxl.load_workbook(src)
ws = wb['Структура сайта']
hdrs = ['URL конкурента 1 (Я.Мск, органика)', 'URL конкурента 2', 'URL конкурента 3', 'Тип выдачи']
start = ws.max_column + 1
for i, h in enumerate(hdrs):
    c = ws.cell(row=1, column=start+i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='2F5496')
    c.alignment = Alignment(horizontal='center', wrap_text=True)

hdr = [(ws.cell(row=1, column=c).value or '') for c in range(1, start)]
url_col = hdr.index('URL') + 1
filled = 0
for r in range(2, ws.max_row+1):
    u = ws.cell(row=r, column=url_col).value
    if u in data:
        ks = data[u]['k'] + ['', '', '']
        for i in range(3):
            ws.cell(row=r, column=start+i, value=ks[i])
        ws.cell(row=r, column=start+3, value=data[u].get('t', ''))
        filled += 1
for i in range(4):
    ws.column_dimensions[openpyxl.utils.get_column_letter(start+i)].width = (46 if i < 3 else 34)
out = os.path.join(HERE, 'struktura-konkurenty-url.xlsx')
wb.save(out)
print('filled', filled, 'of', ws.max_row-1, '->', out)
